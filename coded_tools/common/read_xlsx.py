"""
coded_tools/common/read_xlsx.py
â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
ReadXlsx â€” read data from an Excel (.xlsx) file.

Returns sheet data as CSV-formatted text so agents can parse it easily.
Each row is a comma-separated line; the first row is the header.

Supports:
  â€¢ Selecting a sheet by name, 1-based index, or "list" to see all sheets.
  â€¢ Row-range slicing (start_row / max_rows) for large sheets.
  â€¢ When truncated, the response tells the agent exactly what to call next.

Path resolution:
  â€¢ Relative paths â†’ resolved against input_dir (sly_data)
  â€¢ Absolute paths â†’ used as-is (Flask upload folders)

HOCON class reference
â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    "class": "coded_tools.common.read_xlsx.ReadXlsx"

HOCON tool block (copy-paste into any network)
â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    {
        "name": "read_xlsx",
        "class": "coded_tools.common.read_xlsx.ReadXlsx",
        "function": {
            "description": "Read data from an Excel (.xlsx) file as CSV-formatted text. Use 'sheet' to select a sheet; 'start_row'/'max_rows' for large sheets â€” the response tells you the next start_row when truncated.",
            "parameters": {
                "type": "object",
                "properties": {
                    "path":      { "type": "string",  "description": "XLSX path. Relative to input_dir, or absolute." },
                    "sheet":     { "type": "string",  "description": "Sheet name, 1-based index, or 'list' to see all sheet names (default: first sheet)." },
                    "start_row": { "type": "integer", "description": "1-based data row to start from (optional). Row 1 is the header." },
                    "max_rows":  { "type": "integer", "description": "Maximum number of rows to return per call (default: 500)." },
                    "agent":     { "type": "string",  "description": "Calling agent name (audit log)." }
                },
                "required": ["path"]
            }
        }
    }

sly_data keys read
â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    input_dir      (preferred) â€” where relative paths are resolved
    workspace_dir  (fallback)
    project_folder (fallback)  â€” bidmagic/dealcraft compat
"""

from __future__ import annotations

import asyncio
import csv
import io
import logging
import os
from typing import Any

from neuro_san.interfaces.coded_tool import CodedTool

from coded_tools.common._base import log_call, resolve_input_path

logger = logging.getLogger(__name__)

DEFAULT_MAX_ROWS = 500
MAX_RETURN_BYTES = 64_000


def _cell_value(cell) -> str:
    """Convert an openpyxl cell value to a clean string."""
    if cell.value is None:
        return ""
    return str(cell.value).strip().replace("\n", " ")


class ReadXlsx(CodedTool):
    """Read data from an Excel (.xlsx) file as CSV-formatted text."""

    def invoke(self, args: dict[str, Any], sly_data: dict[str, Any]) -> str:
        try:
            import openpyxl  # type: ignore
        except ImportError:
            return "Error: openpyxl is not installed. Run: pip install openpyxl"

        path_raw  = (args.get("path")  or "").strip()
        agent     = (args.get("agent") or "unknown-agent").strip()
        sheet_id  =  args.get("sheet")
        start_row =  args.get("start_row")
        max_rows  = int(args.get("max_rows") or DEFAULT_MAX_ROWS)

        if not path_raw:
            return "Error: read_xlsx requires 'path'."

        abs_path = resolve_input_path(path_raw, sly_data)

        if not os.path.exists(abs_path):
            log_call(sly_data, tool="ReadXlsx", agent=agent, target=path_raw,
                     status="MISS", detail="file not found")
            return f"NOT_FOUND: '{path_raw}' does not exist."

        if not abs_path.lower().endswith((".xlsx", ".xlsm")):
            return f"Error: '{path_raw}' does not appear to be an .xlsx file."

        try:
            wb = openpyxl.load_workbook(abs_path, read_only=True, data_only=True)
            sheet_names = wb.sheetnames

            # "list" mode â€” return sheet names only
            if str(sheet_id).lower() == "list":
                log_call(sly_data, tool="ReadXlsx", agent=agent, target=path_raw,
                         status="OK", detail=f"listed {len(sheet_names)} sheets")
                return f"Sheets in '{path_raw}':\n" + "\n".join(
                    f"  {i + 1}. {name}" for i, name in enumerate(sheet_names)
                )

            # Select sheet
            if sheet_id is None:
                ws = wb.active
            elif isinstance(sheet_id, int) or (isinstance(sheet_id, str) and sheet_id.isdigit()):
                idx = int(sheet_id) - 1
                if idx < 0 or idx >= len(sheet_names):
                    return (
                        f"Error: sheet index {int(sheet_id)} out of range "
                        f"(workbook has {len(sheet_names)} sheets). "
                        f"Use sheet='list' to see all sheet names."
                    )
                ws = wb[sheet_names[idx]]
            else:
                if str(sheet_id) not in sheet_names:
                    return (
                        f"Error: sheet '{sheet_id}' not found. "
                        f"Available sheets: {', '.join(sheet_names)}"
                    )
                ws = wb[str(sheet_id)]

            sheet_label = ws.title if hasattr(ws, "title") else str(sheet_id or "active")

            # Row-range slicing: start_row is 1-based (row 1 = header)
            s = max(1, int(start_row)) if start_row is not None else 1

            # Build CSV output
            buf = io.StringIO()
            writer = csv.writer(buf, quoting=csv.QUOTE_MINIMAL)
            row_count       = 0   # rows read from sheet (0-based)
            rows_written    = 0   # rows in this slice
            truncated       = False
            total_rows_seen = 0   # total rows in sheet (for hint)

            for row in ws.iter_rows():
                row_count += 1
                total_rows_seen = row_count
                if row_count < s:
                    continue
                if rows_written >= max_rows:
                    # consume remaining to count total (approximated below)
                    truncated = True
                    break
                writer.writerow([_cell_value(c) for c in row])
                rows_written += 1

            wb.close()

        except Exception as exc:
            log_call(sly_data, tool="ReadXlsx", agent=agent, target=path_raw,
                     status="ERROR", detail=str(exc))
            logger.exception("ReadXlsx failed for %s", path_raw)
            return f"Error: failed to read '{path_raw}': {exc}"

        content = buf.getvalue()
        encoded = content.encode("utf-8")

        # Byte-level truncation (fallback if a single row is huge)
        byte_truncated = False
        if len(encoded) > MAX_RETURN_BYTES:
            content = encoded[:MAX_RETURN_BYTES].decode("utf-8", errors="ignore")
            # Find last complete line
            last_nl = content.rfind("\n")
            if last_nl > 0:
                content = content[:last_nl]
            byte_truncated = True
            truncated = True

        sliced = start_row is not None
        detail = f"sheet='{sheet_label}', {rows_written} rows returned"
        if sliced:
            detail += f", starting row {s}"
        if truncated:
            next_s = s + rows_written
            detail += f", truncated (next start_row={next_s})"

        if truncated:
            next_s = s + rows_written
            if byte_truncated:
                hint = (
                    f"\n\n[TRUNCATED â€” output exceeded {MAX_RETURN_BYTES // 1000} KB. "
                    f"Call again with start_row={next_s}"
                    + (f", sheet='{sheet_label}'" if sheet_id else "")
                    + "]"
                )
            else:
                hint = (
                    f"\n\n[TRUNCATED â€” more rows exist. "
                    f"Call again with start_row={next_s}"
                    + (f", sheet='{sheet_label}'" if sheet_id else "")
                    + "]"
                )
            content += hint

        log_call(sly_data, tool="ReadXlsx", agent=agent, target=path_raw,
                 status="OK", detail=detail)
        logger.debug("ReadXlsx: %s  sheet=%s  rows=%d", path_raw, sheet_label, rows_written)
        return content

    async def async_invoke(self, args: dict[str, Any], sly_data: dict[str, Any]) -> str:
        return await asyncio.to_thread(self.invoke, args, sly_data)

