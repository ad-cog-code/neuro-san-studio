"""
coded_tools/common/write_xlsx.py
â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
WriteXlsx â€” write tabular data to an Excel (.xlsx) file.

Accepts content in two formats:
  1. CSV text â€” comma-separated rows, first row treated as header
  2. JSON array of arrays â€” [[col1, col2], [val1, val2], ...]
     or array of objects â€” [{"Name": "Alice", "Score": 95}, ...]

The first row / first object's keys become a bold header row.
Auto-sizes column widths based on content (capped at 60 chars).

Modes:
  â€¢ mode="write"        â€” create new file (or overwrite)
  â€¢ mode="append_sheet" â€” add a new sheet to an existing workbook
  â€¢ mode="append_rows"  â€” add rows to an existing sheet in an existing workbook

HOCON class reference
â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    "class": "coded_tools.common.write_xlsx.WriteXlsx"

HOCON tool block (copy-paste into any network)
â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    {
        "name": "write_xlsx",
        "class": "coded_tools.common.write_xlsx.WriteXlsx",
        "function": {
            "description": "Write tabular data to an Excel (.xlsx) file. Accepts CSV text or JSON array. Creates a new file or appends a new sheet to an existing file.",
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {
                        "type": "string",
                        "description": "Relative path for the output file inside the workspace (e.g. 'outputs/data.xlsx')."
                    },
                    "content": {
                        "type": "string",
                        "description": "Tabular data as CSV text (comma-separated, first row = header) OR as a JSON array of arrays/objects."
                    },
                    "sheet": {
                        "type": "string",
                        "description": "Sheet name to write to (default: 'Sheet1')."
                    },
                    "mode": {
                        "type": "string",
                        "description": "'write' (default â€” create/overwrite file), 'append_sheet' (add a new sheet to an existing file), or 'append_rows' (add rows to an existing sheet)."
                    },
                    "agent": {
                        "type": "string",
                        "description": "Name of the calling agent (used in audit log)."
                    }
                },
                "required": ["path", "content"]
            }
        }
    }

sly_data keys read
â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    output_dir     (preferred) â€” where relative output paths are resolved
    workspace_dir  (fallback)
    project_folder (fallback)  â€” bidmagic/dealcraft compat
"""

from __future__ import annotations

import csv
import io
import json
import logging
import os
from typing import Any

from neuro_san.interfaces.coded_tool import CodedTool

from coded_tools.common._base import log_call, resolve_output_path

logger = logging.getLogger(__name__)

MAX_COL_WIDTH = 60


def _parse_content(content: str) -> list[list[str]]:
    """Parse CSV text or JSON into a list of rows (each row = list of str)."""
    stripped = content.strip()

    # Try JSON first
    if stripped.startswith("["):
        try:
            data = json.loads(stripped)
            if not data:
                return []
            # Array of objects â†’ extract keys as header
            if isinstance(data[0], dict):
                headers = list(data[0].keys())
                rows = [headers]
                for obj in data:
                    rows.append([str(obj.get(h, "")) for h in headers])
                return rows
            # Array of arrays
            return [[str(cell) for cell in row] for row in data]
        except json.JSONDecodeError:
            pass

    # Fall back to CSV parsing
    reader = csv.reader(io.StringIO(stripped))
    return [row for row in reader]


def _auto_width(ws, rows: list[list[str]]):
    """Set column widths based on max content length (capped)."""
    from openpyxl.utils import get_column_letter  # type: ignore
    if not rows:
        return
    col_count = max(len(r) for r in rows)
    for col_idx in range(1, col_count + 1):
        max_len = 0
        for row in rows:
            if col_idx <= len(row):
                max_len = max(max_len, len(str(row[col_idx - 1])))
        width = min(max_len + 2, MAX_COL_WIDTH)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


class WriteXlsx(CodedTool):
    """Write tabular data (CSV or JSON) to an Excel (.xlsx) file."""

    async def async_invoke(self, args: dict[str, Any], sly_data: dict[str, Any]) -> str:
        try:
            import openpyxl                             # type: ignore
            from openpyxl.styles import Font, PatternFill, Alignment  # type: ignore
        except ImportError:
            return "Error: openpyxl is not installed. Run: pip install openpyxl"

        path_raw  = (args.get("path")   or "").strip()
        content   =  args.get("content")
        sheet_name = (args.get("sheet") or "Sheet1").strip()
        mode_raw  = (args.get("mode")   or "write").strip().lower()
        agent     = (args.get("agent")  or "unknown-agent").strip()

        if not path_raw:
            return "Error: write_xlsx requires 'path'."
        if content is None:
            return "Error: write_xlsx requires 'content'."
        if mode_raw not in ("write", "append_sheet", "append_rows"):
            return f"Error: 'mode' must be 'write', 'append_sheet', or 'append_rows' (got '{mode_raw}')."
        if not path_raw.lower().endswith((".xlsx", ".xlsm")):
            path_raw = path_raw + ".xlsx"

        try:
            abs_path = resolve_output_path(path_raw, sly_data)
        except ValueError as exc:
            log_call(sly_data, tool="WriteXlsx", agent=agent, target=path_raw,
                     status="ERROR", detail=str(exc))
            return f"Error: {exc}"

        # Parse content
        try:
            rows = _parse_content(content)
        except Exception as exc:
            return f"Error: could not parse content as CSV or JSON: {exc}"

        if not rows:
            return "Error: content produced no rows to write."

        try:
            os.makedirs(os.path.dirname(abs_path), exist_ok=True)

            # Load existing or create new workbook
            if mode_raw == "append_sheet" and os.path.exists(abs_path):
                wb = openpyxl.load_workbook(abs_path)
                # Ensure unique sheet name
                base = sheet_name
                counter = 1
                while sheet_name in wb.sheetnames:
                    sheet_name = f"{base}_{counter}"
                    counter += 1
                ws = wb.create_sheet(title=sheet_name)
            elif mode_raw == "append_rows" and os.path.exists(abs_path):
                wb = openpyxl.load_workbook(abs_path)
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                else:
                    ws = wb.create_sheet(title=sheet_name)
                # Skip header row when appending rows to an existing sheet with data
                if ws.max_row > 0:
                    rows = rows[1:]   # drop the header â€” it already exists
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = sheet_name

            # Write rows â€” for append_rows, start after existing data
            header_fill = PatternFill(start_color="0033A0", end_color="0033A0", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            row_offset = ws.max_row if mode_raw == "append_rows" and ws.max_row > 0 else 0

            for r_idx, row in enumerate(rows, start=1):
                abs_r = r_idx + row_offset
                for c_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=abs_r, column=c_idx, value=value)
                    # Only format as header when writing the first row in write/append_sheet mode
                    if r_idx == 1 and mode_raw != "append_rows":
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center")

            if mode_raw != "append_rows":
                _auto_width(ws, rows)
            wb.save(abs_path)
            file_size = os.path.getsize(abs_path)

        except Exception as exc:
            log_call(sly_data, tool="WriteXlsx", agent=agent, target=path_raw,
                     status="ERROR", detail=str(exc))
            logger.exception("WriteXlsx failed for %s", path_raw)
            return f"Error: failed to write '{path_raw}': {exc}"

        detail = f"mode={mode_raw}, sheet='{sheet_name}', {len(rows)} rows, {file_size} bytes"
        log_call(sly_data, tool="WriteXlsx", agent=agent, target=path_raw,
                 status="OK", detail=detail)
        logger.debug("WriteXlsx: %s  sheet=%s  rows=%d  size=%dB",
                    path_raw, sheet_name, len(rows), file_size)

        if mode_raw == "append_sheet":
            verb = "appended sheet to"
        elif mode_raw == "append_rows":
            verb = "appended rows to"
        else:
            verb = "wrote"
        return (
            f"OK: {verb} '{path_raw}' â€” "
            f"sheet='{sheet_name}', {len(rows)} rows ({file_size} bytes)."
        )

